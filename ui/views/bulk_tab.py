"""
ui/views/bulk_tab.py — Bulk Excel processing tab
Mendukung multi-worker: setiap worker punya credentials sendiri,
login sendiri, dan memproses partisi baris Excel secara paralel.
"""

import os
import time
import threading
import openpyxl
import tkinter as tk
from tkinter import ttk, messagebox
import customtkinter as ctk

from ui.constants import (WIDGET, ROW_ALT, SUCCESS, DANGER, SUBTEXT, PANEL,
                           TEXT, BORDER, ACCENT_D)
from ui.constants import BTN_ACCENT, BTN_DANGER, BTN_GHOST, BTN_BLUE, BTN_SUBTLE
from ui.constants import FONT, FONT_SMALL, FONT_BOLD
from ui.api import call_bot_api, selenium_login_with_creds
from ui.views.session_bar import _short_login_error

CHECKPOINT = 10


# ── Worker Model ──────────────────────────────────────────────────────────────

class WorkerConfig:
    def __init__(self, label: str, username: str, password: str):
        self.label      = label
        self.username   = username
        self.password   = password
        self.cookie_str = ""
        self.view_state = ""
        self.driver     = None   # live driver — dipakai untuk logout saat card dihapus
        # UI refs (set saat card dibuat)
        self.status_var:   tk.StringVar | None  = None
        self.progress_var: tk.StringVar | None  = None
        self.dot_label:    ctk.CTkLabel | None  = None
        self.login_btn:    ctk.CTkButton | None = None

    @property
    def is_ready(self) -> bool:
        return bool(self.cookie_str and self.view_state)


# ── Main Tab ──────────────────────────────────────────────────────────────────

class BulkTab(ctk.CTkFrame):

    def __init__(self, parent, app):
        super().__init__(parent, fg_color="transparent")
        self._app             = app
        self._excel_rows      = []
        self._progress_max    = 1
        self._last_frac       = 0.0
        self._bulk_start_time = 0.0
        self._workers: list[WorkerConfig] = []   # additional workers
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)
        self._build()

    # ── UI Construction ──────────────────────────────────────────────────────

    def _build(self):
        self._build_toolbar()
        self._build_worker_panel()
        self._build_treeview()

    def _build_toolbar(self):
        toolbar = ctk.CTkFrame(self, fg_color="transparent")
        toolbar.grid(row=0, column=0, sticky="ew", pady=(10, 4), padx=14)

        ctk.CTkLabel(toolbar, text="Search:",
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT).pack(side="left", padx=(0, 4))

        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._on_search())
        ctk.CTkEntry(toolbar, textvariable=self._search_var, width=200,
                     fg_color=WIDGET, border_color=BORDER, border_width=1,
                     text_color=TEXT, font=FONT,
                     placeholder_text="ketik untuk filter...").pack(side="left", padx=(0, 14))

        self._reload_btn = ctk.CTkButton(toolbar, text="↺  Refresh Data Excel",
                                         command=self.load_excel, **BTN_GHOST)
        self._reload_btn.pack(side="left", padx=(0, 6))

        self._run_btn = ctk.CTkButton(toolbar, text="▷ Start",
                                      command=self._run, **BTN_ACCENT)
        self._run_btn.pack(side="left")

        self._stop_btn = ctk.CTkButton(toolbar, text="⏹  Stop",
                                       command=self._stop, state="disabled",
                                       **BTN_DANGER)
        self._stop_btn.pack(side="left", padx=(6, 0))

        ctk.CTkLabel(toolbar, text="Mulai dari baris:",
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT).pack(side="left", padx=(16, 4))

        self._start_from_var = tk.StringVar(value="1")
        self._start_from_cb  = ctk.CTkEntry(toolbar, textvariable=self._start_from_var,
                                             width=64, fg_color=WIDGET,
                                             border_color=BORDER, border_width=1,
                                             text_color=TEXT, font=FONT, justify="center")
        self._start_from_cb.pack(side="left")

        # Tombol buka data.xlsx — float kanan
        ctk.CTkButton(toolbar, text="📂  Buka Data Excel",
                      command=self._open_excel_file,
                      **BTN_BLUE).pack(side="right")

    def _build_worker_panel(self):
        """Strip tipis di bawah toolbar: main worker + scrollable cards + tombol tambah."""
        outer = tk.Frame(self, bg=PANEL, bd=0, highlightthickness=0)
        outer.grid(row=1, column=0, sticky="ew", padx=14, pady=(0, 4))
        # col 2 (canvas) mengisi sisa ruang
        outer.columnconfigure(2, weight=1)

        # ── "Workers :" label ──
        ctk.CTkLabel(outer, text="Workers :",
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).grid(row=0, column=0, padx=(0, 8), pady=1)

        # ── Main worker card ──
        main_card = ctk.CTkFrame(outer, fg_color=WIDGET, corner_radius=6,
                                 border_width=0)
        main_card.grid(row=0, column=1, padx=(0, 6), pady=1)

        self._main_dot = ctk.CTkLabel(main_card, text="●", text_color=SUBTEXT,
                                      fg_color="transparent", font=FONT_SMALL)
        self._main_dot.pack(side="left", padx=(6, 2), pady=1)

        ctk.CTkLabel(main_card, text="Utama",
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).pack(side="left", padx=(0, 2), pady=1)

        self._main_progress_var = tk.StringVar(value="")
        ctk.CTkLabel(main_card, textvariable=self._main_progress_var,
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).pack(side="left", padx=(0, 6), pady=1)

        # ── Scrollable area untuk worker cards tambahan ──
        self._workers_canvas = tk.Canvas(
            outer, bg=PANEL, bd=0, highlightthickness=0, height=30)
        self._workers_canvas.grid(row=0, column=2, sticky="ew", pady=1)

        self._workers_hbar = tk.Scrollbar(
            outer, orient="horizontal",
            command=self._workers_canvas.xview)
        self._workers_hbar.grid(row=1, column=2, sticky="ew")
        self._workers_canvas.configure(xscrollcommand=self._workers_hbar.set)

        self._worker_cards_frame = tk.Frame(
            self._workers_canvas, bg=PANEL, bd=0, highlightthickness=0)
        self._canvas_win = self._workers_canvas.create_window(
            (0, 0), window=self._worker_cards_frame, anchor="nw")

        # Update scroll region setiap kali cards berubah ukuran
        self._worker_cards_frame.bind(
            "<Configure>", self._on_worker_cards_resize)
        # Sesuaikan tinggi canvas mengikuti isi
        self._workers_canvas.bind(
            "<Configure>", lambda e: self._workers_canvas.itemconfig(
                self._canvas_win, height=e.height))

        # Scroll dengan Shift+MouseWheel di atas area cards
        self._workers_canvas.bind(
            "<Shift-MouseWheel>",
            lambda e: self._workers_canvas.xview_scroll(
                -1 if e.delta > 0 else 1, "units"))
        self._worker_cards_frame.bind(
            "<Shift-MouseWheel>",
            lambda e: self._workers_canvas.xview_scroll(
                -1 if e.delta > 0 else 1, "units"))

        # ── Tombol Add Worker ──
        ctk.CTkButton(outer, text="＋  Add Worker",
                      command=self._open_add_worker_dialog,
                      width=120, height=26, **BTN_SUBTLE).grid(
                          row=0, column=3, padx=(8, 0), pady=1)

    def _on_worker_cards_resize(self, _event=None):
        """Perbarui scroll region canvas saat jumlah/ukuran cards berubah."""
        self._workers_canvas.configure(
            scrollregion=self._workers_canvas.bbox("all"))
        # Sembunyikan scrollbar jika tidak diperlukan
        frame_w  = self._worker_cards_frame.winfo_reqwidth()
        canvas_w = self._workers_canvas.winfo_width()
        if frame_w <= canvas_w:
            self._workers_hbar.grid_remove()
        else:
            self._workers_hbar.grid()

    def _build_treeview(self):
        frame = tk.Frame(self, bg=PANEL, highlightthickness=0, bd=0)
        frame.grid(row=2, column=0, sticky="nsew", padx=14, pady=(0, 14))
        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=1)

        cols = ("no", "user_says", "dialog", "score")
        self._tree = ttk.Treeview(frame, columns=cols, show="headings", height=16)
        self._tree.heading("no",        text="#",       anchor="center")
        self._tree.heading("user_says", text="Trigger", anchor="w")
        self._tree.heading("dialog",    text="Dialog",  anchor="w")
        self._tree.heading("score",     text="Score",   anchor="w")
        self._tree.column("no",        width=44,  stretch=False, anchor="center")
        self._tree.column("user_says", width=220, stretch=True)
        self._tree.column("dialog",    width=220, stretch=True)
        self._tree.column("score",     width=220, stretch=True, anchor="w")

        self._tree.tag_configure("odd",       background=WIDGET)
        self._tree.tag_configure("even",      background=ROW_ALT)
        self._tree.tag_configure("done_odd",  background=WIDGET,  foreground=SUCCESS)
        self._tree.tag_configure("done_even", background=ROW_ALT, foreground=SUCCESS)

        vsb = ttk.Scrollbar(frame, orient="vertical",   command=self._tree.yview)
        hsb = ttk.Scrollbar(frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self._tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

    # ── Worker Cards ─────────────────────────────────────────────────────────

    def _open_add_worker_dialog(self):
        WorkerAddDialog(self._app, self._on_worker_added)

    def _on_worker_added(self, username: str, password: str):
        idx = len(self._workers) + 2   # worker 1 = main
        w   = WorkerConfig(label=f"Worker {idx}", username=username, password=password)
        self._workers.append(w)
        self._add_worker_card(w)

    def _bind_scroll(self, widget):
        """Propagate Shift+MouseWheel ke canvas scroll untuk semua child card."""
        widget.bind("<Shift-MouseWheel>",
                    lambda e: self._workers_canvas.xview_scroll(
                        -1 if e.delta > 0 else 1, "units"))
        for child in widget.winfo_children():
            self._bind_scroll(child)

    def _add_worker_card(self, w: WorkerConfig):
        card = ctk.CTkFrame(self._worker_cards_frame, fg_color=WIDGET,
                            corner_radius=6, border_width=0)
        card.pack(side="left", padx=(0, 6))

        w.dot_label = ctk.CTkLabel(card, text="●", text_color=SUBTEXT,
                                   fg_color="transparent", font=FONT_SMALL)
        w.dot_label.pack(side="left", padx=(6, 2), pady=2)

        ctk.CTkLabel(card,
                     text=w.username[:16] + ("…" if len(w.username) > 16 else ""),
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).pack(side="left", padx=(0, 4), pady=2)

        # Status + progress dalam satu var
        w.status_var   = tk.StringVar(value="Belum login")
        w.progress_var = tk.StringVar(value="")
        ctk.CTkLabel(card, textvariable=w.status_var,
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).pack(side="left", pady=2)
        ctk.CTkLabel(card, textvariable=w.progress_var,
                     fg_color="transparent", text_color=SUBTEXT,
                     font=FONT_SMALL).pack(side="left", padx=(2, 4), pady=2)

        # Tombol Login manual
        w.login_btn = ctk.CTkButton(card, text="Login", width=48, height=20,
                                    font=FONT_SMALL,
                                    fg_color="transparent", hover_color=BORDER,
                                    text_color=SUBTEXT, border_width=1,
                                    border_color=BORDER,
                                    command=lambda wk=w: self._login_worker(wk))
        w.login_btn.pack(side="left", padx=(0, 4), pady=2)

        ctk.CTkButton(card, text="✕", width=20, height=20,
                      fg_color="transparent", hover_color=BORDER,
                      text_color=SUBTEXT, font=FONT_SMALL,
                      command=lambda: self._remove_worker(w, card)).pack(
                          side="left", padx=(0, 4))

        # Bind scroll setelah semua widget dalam card terbentuk
        self._app.after(10, self._bind_scroll, card)

    def _remove_worker(self, w: WorkerConfig, card):
        if w in self._workers:
            self._workers.remove(w)
        card.destroy()
        if w.driver:
            threading.Thread(target=self._logout_worker_driver,
                             args=(w,), daemon=True).start()

    @staticmethod
    def _logout_worker_driver(w: WorkerConfig):
        """Logout server-side lalu quit driver. Jalan di background thread."""
        driver = w.driver
        w.driver = None
        if not driver:
            return
        try:
            from pages.dashboard_page import DashboardPage
            DashboardPage(driver).logout()
        except Exception:
            pass
        finally:
            try:
                driver.quit()
            except Exception:
                pass

    def logout_all_workers(self):
        """Logout semua worker yang masih punya driver. Dipanggil saat app ditutup."""
        workers_with_driver = [w for w in self._workers if w.driver]
        if not workers_with_driver:
            return
        threads = [
            threading.Thread(target=self._logout_worker_driver, args=(w,), daemon=True)
            for w in workers_with_driver
        ]
        for t in threads:
            t.start()
        for t in threads:
            t.join(timeout=20)   # tunggu max 20 detik per batch

    def _login_worker(self, w: WorkerConfig):
        """Login manual satu worker — bisa dilakukan sebelum klik Mulai."""
        if w.login_btn:
            w.login_btn.configure(state="disabled")
        self._set_worker_status(w, "Login...", SUBTEXT)

        def _on_progress(_):
            pass  # sudah cukup indikasi dari status_var

        def _on_done(driver, cookie_str, view_state):
            w.driver     = driver   # simpan — dipakai logout saat card dihapus
            w.cookie_str = cookie_str
            w.view_state = view_state
            self._set_worker_status(w, "✓ Ready", SUCCESS)
            if w.login_btn:
                self._app.after(0, w.login_btn.configure, {"state": "disabled", "text": "✓"})

        def _on_error(msg):
            self._set_worker_status(w, "Gagal", DANGER)
            if w.login_btn:
                self._app.after(0, w.login_btn.configure, {"state": "normal"})
            self._app.after(0, messagebox.showerror,
                            "Worker Login Error",
                            f"{w.label} ({w.username}):\n{_short_login_error(msg)}")

        threading.Thread(
            target=selenium_login_with_creds,
            args=(w.username, w.password, _on_progress, _on_done, _on_error),
            daemon=True,
        ).start()

    def _set_worker_status(self, w: WorkerConfig, status: str, color: str):
        """Thread-safe update status card worker."""
        def _do():
            if w.status_var:
                w.status_var.set(status)
            if w.dot_label:
                w.dot_label.configure(text_color=color)
        self._app.after(0, _do)

    # ── Excel Load ───────────────────────────────────────────────────────────

    def _open_excel_file(self):
        path = self._app.EXCEL_PATH
        if os.path.exists(path):
            os.startfile(path)
        else:
            messagebox.showwarning("File tidak ditemukan",
                                   "data.xlsx belum ada di direktori project.")

    def load_excel(self):
        excel_path = self._app.EXCEL_PATH
        if not os.path.exists(excel_path):
            self._app.session_bar.update_bulk(0, "data.xlsx tidak ditemukan di direktori project.")
            return
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            self._search_var.set("")
            self._tree.delete(*self._tree.get_children())
            self._excel_rows.clear()
            seq = 0
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=1).value
                if not val:
                    continue
                seq += 1
                dialog = ws.cell(row=row, column=2).value or ""
                score  = ws.cell(row=row, column=3).value or ""
                tag    = "even" if seq % 2 == 0 else "odd"
                iid = self._tree.insert("", "end",
                                        values=(seq, str(val), str(dialog), str(score)),
                                        tags=(tag,))
                self._excel_rows.append((row, str(val), iid))
            total = len(self._excel_rows)
            self._app.session_bar.update_bulk(0, f"Berhasil Memuat {total} Baris Trigger")
            self._start_from_var.set("1")
        except Exception as exc:
            messagebox.showerror("Load Error", str(exc))

    # ── Bulk Run ─────────────────────────────────────────────────────────────

    def _run(self):
        app = self._app
        if not app.check_ready():
            return
        if not self._excel_rows:
            messagebox.showwarning("Kosong", "Tidak ada data. Klik 'Reload' terlebih dahulu.")
            return

        try:
            start_from = max(1, int(self._start_from_var.get() or "1"))
        except ValueError:
            start_from = 1

        app._cancel = False
        self._run_btn.configure(state="disabled")
        self._reload_btn.configure(state="disabled")
        self._start_from_cb.configure(state="disabled")
        self._stop_btn.configure(state="normal")
        self._main_progress_var.set("")
        self._main_dot.configure(text_color=SUBTEXT)

        excel_rows = list(self._excel_rows)[start_from - 1:]
        total      = len(excel_rows)
        self._progress_max    = max(total, 1)
        self._last_frac       = 0.0
        self._bulk_start_time = time.time()
        self._app.session_bar.update_bulk(0, f"0 / {total}")

        all_workers_count = 1 + len(self._workers)
        partitions        = _partition(excel_rows, all_workers_count)
        main_session      = (app._cookie_str, app._view_state)

        results      : dict           = {}
        results_lock : threading.Lock = threading.Lock()
        progress_state                = {"done": 0, "total": total}

        def _on_all_done():
            self._write_results_to_excel(results, excel_rows)
            self._finish()

        # ── helper: jalankan semua partisi paralel setelah semua worker ready ──
        def _start_all_partitions():
            # Progress callback main worker
            def _main_prog(done_n, part_total):
                self._app.after(0, self._main_progress_var.set,
                                f"{done_n}/{part_total}")

            threads = [threading.Thread(
                target=self._run_partition,
                args=(main_session, partitions[0], results, results_lock,
                      progress_state, None, _main_prog),
                daemon=True,
            )]

            for i, w in enumerate(self._workers):
                w_ref = w
                def _make_prog(worker):
                    def _wp(done_n, part_total):
                        self._app.after(0, worker.progress_var.set,
                                        f"{done_n}/{part_total}")
                    return _wp

                threads.append(threading.Thread(
                    target=self._run_partition,
                    args=((w_ref.cookie_str, w_ref.view_state),
                          partitions[i + 1], results, results_lock,
                          progress_state, None, _make_prog(w_ref)),
                    daemon=True,
                ))
                self._set_worker_status(w_ref, "▶ Jalan", SUBTEXT)

            self._app.after(0, self._main_dot.configure, {"text_color": SUBTEXT})

            def _wait_all():
                for t in threads:
                    t.join()
                self._app.after(0, _on_all_done)

            for t in threads:
                t.start()
            threading.Thread(target=_wait_all, daemon=True).start()

        # ── jika tidak ada additional workers langsung mulai ──────────────────
        if not self._workers:
            threading.Thread(target=_start_all_partitions, daemon=True).start()
            return

        # ── login worker yang belum ready, skip yang sudah ready ─────────────
        ready_lock   = threading.Lock()
        ready_count  = [0]
        login_failed = [False]
        need_login = [w for w in self._workers if not w.is_ready]

        # Hitung total yang harus "check in" (main + need_login workers)
        # Workers yang sudah ready tidak perlu login, langsung increment
        all_count = 1 + len(need_login)

        def _maybe_start():
            with ready_lock:
                ready_count[0] += 1
                if ready_count[0] < all_count:
                    return
            if login_failed[0]:
                self._app.after(0, self._finish)
                return
            _start_all_partitions()

        # Main worker langsung check in
        threading.Thread(target=_maybe_start, daemon=True).start()

        for w in need_login:
            self._set_worker_status(w, "Login...", SUBTEXT)
            if w.login_btn:
                self._app.after(0, w.login_btn.configure, {"state": "disabled"})

            def _make_callbacks(worker):
                def _on_progress(_):
                    pass

                def _on_done(driver, cookie_str, view_state):
                    worker.driver     = driver   # simpan untuk logout nanti
                    worker.cookie_str = cookie_str
                    worker.view_state = view_state
                    self._set_worker_status(worker, "✓ Ready", SUCCESS)
                    _maybe_start()

                def _on_error(msg):
                    login_failed[0] = True
                    self._set_worker_status(worker, "Gagal", DANGER)
                    self._app.after(0, messagebox.showerror,
                                    "Worker Login Error",
                                    f"{worker.label} ({worker.username}):\n{_short_login_error(msg)}")
                    _maybe_start()

                return _on_progress, _on_done, _on_error

            prog, done_cb, err = _make_callbacks(w)
            threading.Thread(
                target=selenium_login_with_creds,
                args=(w.username, w.password, prog, done_cb, err),
                daemon=True,
            ).start()

    def _run_partition(self, session: tuple, rows: list,
                       results: dict, lock: threading.Lock,
                       progress_state: dict, on_done_cb,
                       worker_progress_cb=None):
        """Proses satu partisi baris. worker_progress_cb(done, total) dipanggil per baris."""
        cookie_str, view_state = session
        app        = self._app
        part_done  = 0
        part_total = len(rows)

        for row_num, user_says, iid in rows:
            if app._cancel:
                break
            try:
                for reset_word in ("menu", "batal"):
                    call_bot_api(cookie_str, view_state, reset_word)
                result = call_bot_api(cookie_str, view_state, user_says)
            except Exception as e:
                result = {"dialog": f"[ERROR] {e}", "score": ""}

            part_done += 1
            with lock:
                results[row_num] = (result["dialog"], result["score"])
                progress_state["done"] += 1
                global_done  = progress_state["done"]
                global_total = progress_state["total"]

            app.after(0, self._update_row, iid,
                      result["dialog"], result["score"],
                      global_done, global_total)

            if worker_progress_cb:
                worker_progress_cb(part_done, part_total)

        if on_done_cb:
            app.after(0, on_done_cb)

    def _write_results_to_excel(self, results: dict, excel_rows: list):
        """
        Tulis hasil ke Excel dengan file lock cross-process.
        Hanya update baris yang diproses instance ini — instance lain tidak
        ikut tertimpa meskipun berjalan bersamaan.
        """
        excel_path = self._app.EXCEL_PATH
        lock_path  = excel_path + ".lock"
        import time as _time

        # Tunggu lock dilepas instance lain (max 60 detik)
        waited = 0
        while os.path.exists(lock_path) and waited < 60:
            _time.sleep(0.5)
            waited += 0.5

        # Ambil lock
        try:
            with open(lock_path, "w") as lf:
                lf.write(str(os.getpid()))
        except Exception:
            pass

        try:
            # Load file segar — bukan dari memori awal sesi
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            ws["B1"] = "Dialog"
            ws["C1"] = "Confident Score"
            # Hanya update baris milik instance ini
            for row_num, _user_says, _iid in excel_rows:
                if row_num in results:
                    dialog, score = results[row_num]
                    ws.cell(row=row_num, column=2).value = dialog
                    ws.cell(row=row_num, column=3).value = score
            wb.save(excel_path)
            saved   = len(results)
            skipped = len(excel_rows) - saved
            if self._app._cancel and skipped > 0:
                msg = f"⏹  Dihentikan. {saved} baris tersimpan, {skipped} dilewati."
            else:
                msg = f"✓  Selesai! {saved} baris disimpan ke data.xlsx"
            self._app.after(0, self._app.session_bar.update_bulk,
                            saved / max(len(excel_rows), 1), msg)
        except PermissionError:
            self._app.after(0, messagebox.showerror, "Simpan Gagal",
                            "Tutup file data.xlsx di Excel terlebih dahulu.")
        except Exception as exc:
            self._app.after(0, messagebox.showerror, "Simpan Gagal", str(exc))
        finally:
            # Lepas lock
            try:
                os.remove(lock_path)
            except Exception:
                pass

    # ── Helpers ──────────────────────────────────────────────────────────────

    def _update_row(self, iid, dialog, score, i, total):
        try:
            vals     = self._tree.item(iid, "values")
            old_tags = self._tree.item(iid, "tags")
            done_tag = "done_even" if "even" in old_tags else "done_odd"
            self._tree.item(iid, values=(vals[0], vals[1], dialog, score),
                            tags=(done_tag,))
        except tk.TclError:
            pass
        self._last_frac = i / self._progress_max
        self._app.session_bar.update_bulk(self._last_frac, f"{i} / {total}")

        # ETA di subtitle kiri bawah
        if i > 0 and i < total and self._bulk_start_time:
            elapsed  = time.time() - self._bulk_start_time
            remaining = (elapsed / i) * (total - i)
            self._app.session_bar.set_progress(_fmt_eta(remaining))
        elif i >= total:
            self._app.session_bar.set_progress("")

    def _on_search(self):
        q = self._search_var.get().lower().strip()
        if not q:
            self._tree.selection_set([])
            return
        matches = [
            iid for iid in self._tree.get_children()
            if any(q in str(v).lower() for v in self._tree.item(iid, "values")[1:])
        ]
        self._tree.selection_set(matches)
        if matches:
            self._tree.see(matches[0])

    def _stop(self):
        self._app._cancel = True
        self._stop_btn.configure(state="disabled")

    def _finish(self):
        self._run_btn.configure(state="normal")
        self._reload_btn.configure(state="normal")
        self._start_from_cb.configure(state="normal")
        self._stop_btn.configure(state="disabled")
        self._main_progress_var.set("")
        self._main_dot.configure(text_color=SUCCESS)
        self._bulk_start_time = 0.0
        self._app.session_bar.set_progress("")
        for w in self._workers:
            if w.is_ready:
                self._set_worker_status(w, "✓ Selesai", SUCCESS)
                if w.progress_var:
                    self._app.after(0, w.progress_var.set, "")


# ── ETA helper ───────────────────────────────────────────────────────────────

def _fmt_eta(seconds: float) -> str:
    """Format sisa waktu menjadi string singkat, misal '~2m 30d' atau '~45d'."""
    s = max(0, int(seconds))
    if s < 60:
        return f"›  Estimasi selesai: ~{s} detik lagi"
    m, s = divmod(s, 60)
    if m < 60:
        return f"›  Estimasi selesai: ~{m}m {s:02d}d lagi"
    h, m = divmod(m, 60)
    return f"›  Estimasi selesai: ~{h}j {m}m lagi"


# ── Partition helper ──────────────────────────────────────────────────────────

def _partition(lst: list, n: int) -> list[list]:
    """Bagi lst menjadi n partisi kontiguous sebisa mungkin merata."""
    if n <= 1:
        return [lst]
    size  = len(lst)
    chunk = max(1, (size + n - 1) // n)
    return [lst[i:i + chunk] for i in range(0, size, chunk)]


# ── Worker Add Dialog ─────────────────────────────────────────────────────────

class WorkerAddDialog(ctk.CTkToplevel):

    def __init__(self, parent, on_save):
        super().__init__(parent)
        self._on_save = on_save
        self.title("Tambah Worker")
        self.resizable(False, False)
        self.grab_set()
        self._build()
        self.after(50, self._center)

    def _center(self):
        self.update_idletasks()
        pw = self.master.winfo_rootx()
        py = self.master.winfo_rooty()
        pw2 = self.master.winfo_width()
        py2 = self.master.winfo_height()
        w, h = self.winfo_width(), self.winfo_height()
        self.geometry(f"+{pw + (pw2 - w)//2}+{py + (py2 - h)//2}")

    def _build(self):
        pad = {"padx": 20, "pady": (10, 4)}
        ctk.CTkLabel(self, text="Username / Email", font=FONT).pack(**pad, anchor="w")
        self._user_entry = ctk.CTkEntry(self, width=300, font=FONT)
        self._user_entry.pack(padx=20, pady=(0, 10))

        ctk.CTkLabel(self, text="Password", font=FONT).pack(**pad, anchor="w")
        self._pass_entry = ctk.CTkEntry(self, width=300, font=FONT, show="●")
        self._pass_entry.pack(padx=20, pady=(0, 16))

        btns = ctk.CTkFrame(self, fg_color="transparent")
        btns.pack(padx=20, pady=(0, 16), fill="x")
        ctk.CTkButton(btns, text="Batal", command=self.destroy,
                      **BTN_GHOST).pack(side="left")
        ctk.CTkButton(btns, text="Tambah", command=self._save,
                      **BTN_ACCENT).pack(side="right")

    def _save(self):
        username = self._user_entry.get().strip()
        password = self._pass_entry.get().strip()
        if not username or not password:
            messagebox.showwarning("Kosong", "Username dan password wajib diisi.",
                                   parent=self)
            return
        self._on_save(username, password)
        self.destroy()
